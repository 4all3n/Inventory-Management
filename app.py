from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
import openpyxl


app = Flask(__name__)
app.secret_key = 'your_secret_key'

EXCEL_FILE = "data/inventory_data.xlsx"

def initialize_excel_file(filepath):
    required_sheets = {
        "inventory": ['HSN_Code', 'Product_Name', 'Units', 'Cost', 'Selling_Price','Total_cost'],
        "products": ['HSN_Code', 'Product_Name', 'Cost', 'Selling_Price'],
        "customers": ['Customer_ID', 'Name', 'Email', 'Address'],
        "vendors": ['HSN_Code', 'Vendor_Name', 'Product_Name', 'Phone', 'Email', 'Address'],
        "purchases": ['HSN_Code', 'Product_Name', 'Vendor', 'Date', 'Units', 'Cost_per_Unit', 'Total_Cost'],
        "sales": ['HSN_Code', 'Product_Name', 'Customer', 'Date', 'Units', 'Selling_Price_per_Unit', 'Total_Amount'],
    }

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if not os.path.exists(filepath):
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for sheet, headers in required_sheets.items():
                df = pd.DataFrame(columns=headers)
                df.to_excel(writer, sheet_name=sheet, index=False)
        print("✅ Excel file and sheets created.")
    else:
        workbook = load_workbook(filepath)
        existing_sheets = workbook.sheetnames
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            for sheet, headers in required_sheets.items():
                if sheet not in existing_sheets:
                    df = pd.DataFrame(columns=headers)
                    df.to_excel(writer, sheet_name=sheet, index=False)
                    print(f"✅ Sheet '{sheet}' created.")

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/add_customer', methods=['GET', 'POST'])
def add_customer_route():
    if request.method == 'POST':
        cust_id = request.form['cust_id']
        name = request.form['name']
        email = request.form['email']
        address = request.form['address']

        try:
            df_customer = pd.read_excel(EXCEL_FILE, sheet_name='customers')
        except ValueError:
            df_customer = pd.DataFrame(columns=['Customer_ID', 'Name', 'Email', 'Address'])

        if cust_id in df_customer['Customer_ID'].astype(str).values:
            flash("Customer ID already exists. Please use a different ID.")
            return redirect(url_for('add_customer_route'))

        new_customer = {
            'Customer_ID': cust_id,
            'Name': name,
            'Email': email,
            'Address': address
        }

        df_customer = pd.concat([df_customer, pd.DataFrame([new_customer])], ignore_index=True)

        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_customer.to_excel(writer, sheet_name='customers', index=False)

        flash("Customer added successfully!")
        return redirect(url_for('add_customer_route'))

    return render_template('add_customer.html')

@app.route('/add_product', methods=['GET', 'POST'])
def add_product_route():
    if request.method == 'POST':
        hsn_code = request.form.get('hsn_code').strip()
        product_name = request.form.get('product_name').strip()
        cost = request.form.get('cost').strip()
        selling_price = request.form.get('selling_price').strip()

        if not hsn_code or not product_name or not cost or not selling_price:
            flash("All fields are required.", "danger")
            return redirect(url_for('add_product_route'))

        try:
            cost = float(cost)
            selling_price = float(selling_price)

            xls = pd.ExcelFile(EXCEL_FILE)

            df_products = pd.read_excel(xls, sheet_name='products')
            new_product = pd.DataFrame([{
                'HSN_Code': hsn_code,
                'Product_Name': product_name,
                'Cost': cost,
                'Selling_Price': selling_price
            }])
            df_products = pd.concat([df_products, new_product], ignore_index=True)

            df_inventory = pd.read_excel(xls, sheet_name='inventory')
            new_inventory = pd.DataFrame([{
                'HSN_Code': hsn_code,
                'Product_Name': product_name,
                'Units': 0,
                'Cost': cost,
                'Selling_Price': selling_price,
                'Total_cost': 0
            }])
            df_inventory = pd.concat([df_inventory, new_inventory], ignore_index=True)

            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_products.to_excel(writer, sheet_name='products', index=False)
                df_inventory.to_excel(writer, sheet_name='inventory', index=False)

            flash("✅ Product added successfully!", "success")
            return redirect(url_for('add_product_route'))

        except Exception as e:
            print("Error:", e)
            flash("❌ Failed to add product. Please try again.", "danger")
            return redirect(url_for('add_product_route'))

    return render_template('add_product.html')

@app.route('/add_vendor', methods=['GET', 'POST'])
def add_vendor_route():
    df_product = pd.read_excel(EXCEL_FILE, sheet_name='products')
    product_dict = df_product.set_index('HSN_Code')['Product_Name'].to_dict()

    if request.method == 'POST':
        hsn_code = request.form['hsn_code']
        vendor_name = request.form['vendor_name']
        product_name = request.form['product_name']
        phone = request.form['phone']
        email = request.form['email']
        address = request.form['address']

        df_vendor = pd.read_excel(EXCEL_FILE, sheet_name='vendors')
        new_row = {
            'HSN_Code': hsn_code,
            'Vendor_Name': vendor_name,
            'Product_Name': product_name,
            'Phone': phone,
            'Email': email,
            'Address': address
        }
        df_vendor = pd.concat([df_vendor, pd.DataFrame([new_row])], ignore_index=True)

        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_vendor.to_excel(writer, sheet_name='vendors', index=False)

        flash("Vendor added successfully!")
        return redirect(url_for('add_vendor_route'))

    return render_template('add_vendor.html', product_dict=product_dict)

@app.route('/add_purchase', methods=['GET', 'POST'])
def add_purchase():
    vendor_df = pd.read_excel(EXCEL_FILE, sheet_name="vendors")

    hsn_codes = sorted(vendor_df['HSN_Code'].dropna().unique())
    product_names = sorted(vendor_df['Product_Name'].dropna().unique())
    vendors = sorted(vendor_df['Vendor_Name'].dropna().unique())

    # Load inventory sheet
    try:
        inventory_df = pd.read_excel(EXCEL_FILE, sheet_name="inventory")
    except:
        inventory_df = pd.DataFrame(columns=[
            'HSN_Code', 'Product_Name', 'Units', 'Cost', 'Selling_Price', 'Total_cost'
        ])

    hsn_to_product = dict(zip(vendor_df['HSN_Code'].astype(str), vendor_df['Product_Name']))
    hsn_to_selling_price = dict(zip(inventory_df['HSN_Code'].astype(str), inventory_df['Selling_Price'].fillna(0).astype(float)))

    if request.method == 'POST':
        hsn_code = request.form['hsn_code']
        product_name = request.form['product_name']
        vendor = request.form['vendor']
        date = request.form['date']
        units = int(request.form['units'])
        cost_per_unit = float(request.form['cost_per_unit'])
        total_cost = units * cost_per_unit

        try:
            # Read all sheets at once to avoid multiple file operations
            with pd.ExcelFile(EXCEL_FILE) as xls:
                purchase_df = pd.read_excel(xls, sheet_name="purchases")
                inventory_df = pd.read_excel(xls, sheet_name="inventory")

            # Add new purchase
            new_row = pd.DataFrame([{
                'HSN_Code': hsn_code,
                'Product_Name': product_name,
                'Vendor': vendor,
                'Date': date,
                'Units': units,
                'Cost_per_Unit': cost_per_unit,
                'Total_Cost': total_cost
            }])
            purchase_df = pd.concat([purchase_df, new_row], ignore_index=True)

            # Clean values for safer comparison
            hsn_code = str(hsn_code).strip()
            product_name = str(product_name).strip()

            inventory_df['HSN_Code'] = inventory_df['HSN_Code'].astype(str).str.strip()
            inventory_df['Product_Name'] = inventory_df['Product_Name'].astype(str).str.strip()

            match = (
                (inventory_df['HSN_Code'] == hsn_code) &
                (inventory_df['Product_Name'] == product_name)
            )

            if match.any():
                # Update existing inventory entry
                inventory_df.loc[match, 'Units'] = (
                    inventory_df.loc[match, 'Units'].astype(float) + units
                )
                inventory_df.loc[match, 'Cost'] = cost_per_unit
                inventory_df.loc[match, 'Total_cost'] = (
                    inventory_df.loc[match, 'Units'].astype(float) * cost_per_unit
                )
            else:
                # Add new entry
                new_inv = pd.DataFrame([{
                    'HSN_Code': hsn_code,
                    'Product_Name': product_name,
                    'Units': units,
                    'Cost': cost_per_unit,
                    'Selling_Price': 0,  # default
                    'Total_cost': total_cost
                }])
                inventory_df = pd.concat([inventory_df, new_inv], ignore_index=True)

            # Save all changes at once
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                purchase_df.to_excel(writer, sheet_name='purchases', index=False)
                inventory_df.to_excel(writer, sheet_name='inventory', index=False)

            flash("Purchase added successfully!")
            return redirect(url_for('add_purchase'))

        except Exception as e:
            print(f"Error adding purchase: {str(e)}")
            flash("Error adding purchase. Please try again.")
            return redirect(url_for('add_purchase'))

    return render_template('add_purchase.html', 
                         hsn_codes=hsn_codes,
                         product_names=product_names,
                         vendors=vendors,
                         hsn_to_product=hsn_to_product,
                         hsn_to_selling_price=hsn_to_selling_price)



@app.route('/add_sales', methods=['GET', 'POST'])
def add_sales():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    inventory_sheet = wb['inventory']
    sales_sheet = wb['sales']
    customer_sheet = wb['customers']

    # ---- Prepare dropdown mappings ----
    hsn_to_product = {}
    product_to_hsn = {}
    selling_prices = {}
    stocks = {}

    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        hsn, product, stock, _, selling_price, _ = row
        if hsn is None or product is None:
            continue
        hsn = str(hsn)
        hsn_to_product[hsn] = product
        product_to_hsn[product] = hsn
        selling_prices[hsn] = selling_price
        stocks[hsn] = stock

    customers = []
    for row in customer_sheet.iter_rows(min_row=2, values_only=True):
        cust_id, name = row[:2]
        if cust_id and name:
            customers.append(f"{cust_id} - {name}")

    # ---- Handle Form POST ----
    if request.method == 'POST':
        hsn_code = request.form['hsn_code']
        product_name = request.form['product_name']
        customer = request.form['customer']
        date = request.form['date']
        units = int(request.form['units'])
        selling_price = float(request.form['selling_price'])
        total_cost = units * selling_price

        available_stock = stocks.get(hsn_code, 0)

        if units > available_stock:
            flash(f"❌ Not enough stock available. Only {available_stock} units left.", 'error')
            return redirect('/add_sales')

        # ✅ Add entry to Sales sheet (using pandas)
        try:
            sales_df = pd.read_excel(EXCEL_FILE, sheet_name="sales")
        except:
            sales_df = pd.DataFrame(columns=[
                'HSN_Code', 'Product_Name', 'Customer', 'Date',
                'Units', 'Selling_Price_per_Unit', 'Total_Amount'
            ])

        new_row = pd.DataFrame([{
            'HSN_Code': hsn_code,
            'Product_Name': product_name,
            'Customer': customer,
            'Date': date,
            'Units': units,
            'Selling_Price_per_Unit': selling_price,
            'Total_Amount': total_cost
        }])
        sales_df = pd.concat([sales_df, new_row], ignore_index=True)

        # ✅ Update inventory stock using openpyxl
        for row in inventory_sheet.iter_rows(min_row=2):
            cell_hsn = str(row[0].value)
            if cell_hsn == hsn_code:
                row[2].value = row[2].value - units  # Subtract sold units
                break

        # Save both sheets
        wb.save(EXCEL_FILE)
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            sales_df.to_excel(writer, sheet_name="sales", index=False)

        flash("✅ Sale added successfully!")
        return redirect('/add_sales')

    # Render sales form
    return render_template(
        'add_sales.html',
        hsn_to_product=hsn_to_product,
        product_to_hsn=product_to_hsn,
        selling_prices=selling_prices,
        stocks=stocks,
        customers=customers
    )



# Route for viewing all sheets
@app.route('/sheets')
def view_all_sheets():
    sheets = ['customers', 'products', 'vendors', 'purchases', 'sales', 'inventory']
    return render_template('view_all_sheets.html', sheets=sheets)

# Route for viewing individual sheet data
@app.route('/sheets/<sheet_name>')
def view_sheet(sheet_name): 

    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        headers = df.columns.tolist()
        rows = df.values.tolist()
    except Exception as e:
        return f"Error loading sheet '{sheet_name}': {e}", 500

    return render_template('view_sheet.html', sheet_name=sheet_name, headers=headers, rows=rows)

@app.route('/sheets/<sheet_name>/edit/<int:row_index>', methods=['GET', 'POST'])
def edit_row(sheet_name, row_index):
    all_sheets = pd.read_excel(EXCEL_FILE, sheet_name=None)
    df = all_sheets.get(sheet_name)

    if df is None:
        return f"Sheet '{sheet_name}' not found", 404
    
    # Convert 1-based index to 0-based index
    actual_index = row_index - 1
    
    if actual_index < 0 or actual_index >= len(df):
        return f"Invalid row index: {row_index}", 404

    if request.method == 'POST':
        print("Form Data:", request.form)  # Debug print
        inventory_df = all_sheets.get('inventory')

        # Step 1: Fetch old row before update
        old_row = df.iloc[actual_index].copy()
        print("Old Row:", old_row.to_dict())  # Debug print
        old_units = int(old_row.get('Units', 0) or 0)
        old_hsn = str(old_row.get('HSN_Code')).strip()
        old_product = str(old_row.get('Product_Name')).strip()

        # Step 2: Get new form values and convert to appropriate types
        form_data = {}
        for header in df.columns:
            # Get the last value if there are duplicates
            values = request.form.getlist(header)
            value = values[-1] if values else None
            
            print(f"Processing {header}: {value}")  # Debug print
            if value is not None and value != '':
                try:
                    # Convert values based on column type
                    if header in ['Units']:
                        form_data[header] = int(value)
                    elif header in ['Cost_per_Unit', 'Selling_Price_per_Unit', 'Total_Cost', 'Total_Amount']:
                        form_data[header] = float(value)
                    elif header == 'HSN_Code':
                        form_data[header] = str(value).strip()
                    else:
                        form_data[header] = str(value).strip()
                except ValueError as e:
                    print(f"Error converting {header}: {e}")  # Debug print
                    form_data[header] = old_row[header]
            else:
                form_data[header] = old_row[header]

        print("Form Data after processing:", form_data)  # Debug print

        # Step 3: Update the row with new values
        try:
            for header in df.columns:
                df.at[actual_index, header] = form_data[header]
            print("Row updated successfully")  # Debug print
        except Exception as e:
            print(f"Error updating row: {e}")  # Debug print
            flash("Error updating row. Please try again.", "error")
            return redirect(url_for('view_sheet', sheet_name=sheet_name))

        # Step 4: Adjust inventory if needed
        if inventory_df is not None:
            try:
                # Clean up HSN codes and product names for comparison
                inventory_df['HSN_Code'] = inventory_df['HSN_Code'].astype(str).str.strip()
                inventory_df['Product_Name'] = inventory_df['Product_Name'].astype(str).str.strip()

                # Find matching inventory row
                inv_match = (
                    (inventory_df['HSN_Code'] == form_data['HSN_Code']) &
                    (inventory_df['Product_Name'] == form_data['Product_Name'])
                )
                
                if inv_match.any():
                    inv_index = inv_match.idxmax()
                    current_stock = int(inventory_df.at[inv_index, 'Units'])

                    if sheet_name.lower() == 'purchases':
                        # For purchases: new_units - old_units
                        delta = int(form_data['Units']) - old_units
                        inventory_df.at[inv_index, 'Units'] = current_stock + delta
                        # Update cost and total cost
                        inventory_df.at[inv_index, 'Cost'] = float(form_data['Cost_per_Unit'])
                        inventory_df.at[inv_index, 'Total_cost'] = float(form_data['Total_Cost'])
                    elif sheet_name.lower() == 'sales':
                        # For sales: old_units - new_units
                        delta = old_units - int(form_data['Units'])
                        inventory_df.at[inv_index, 'Units'] = current_stock + delta

                    print(f"Inventory updated: {delta} units, new stock: {inventory_df.at[inv_index, 'Units']}")
                else:
                    print("No matching inventory row found")

                all_sheets['inventory'] = inventory_df
            except Exception as e:
                print(f"Error updating inventory: {e}")  # Debug print
                flash("Error updating inventory. Please try again.", "error")
                return redirect(url_for('view_sheet', sheet_name=sheet_name))

        # Step 5: Save all sheets back
        try:
            all_sheets[sheet_name] = df
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='w') as writer:
                for name, data in all_sheets.items():
                    data.to_excel(writer, sheet_name=name, index=False)
            print("Excel file saved successfully")  # Debug print
        except Exception as e:
            print(f"Error saving Excel file: {e}")  # Debug print
            flash("Error saving changes. Please try again.", "error")
            return redirect(url_for('view_sheet', sheet_name=sheet_name))

        flash(f"Row updated successfully!", "success")
        return redirect(url_for('view_sheet', sheet_name=sheet_name))

    # GET: Render edit form
    headers = df.columns.tolist()
    row_data = df.iloc[actual_index].tolist()
    zipped = list(zip(headers, row_data))

    if sheet_name.lower() in ['purchases', 'sales']:
        products_df = all_sheets.get('products')
        vendors_df = all_sheets.get('vendors')

        hsn_to_product = dict(zip(products_df['HSN_Code'], products_df['Product_Name']))
        product_to_hsn = dict(zip(products_df['Product_Name'], products_df['HSN_Code']))
        hsn_codes = list(hsn_to_product.keys())
        product_names = list(hsn_to_product.values())
        vendors = vendors_df['Vendor_Name'].tolist()

        return render_template('edit_purchase_and_sales.html',
                               row_data=row_data,
                               zipped=zipped,
                               sheet_name=sheet_name,
                               hsn_codes=hsn_codes,
                               product_names=product_names,
                               vendors=vendors,
                               hsn_to_product=hsn_to_product,
                               product_to_hsn=product_to_hsn)
    else:
        return render_template('edit_row.html',
                               row_data=row_data,
                               zipped=zipped,
                               sheet_name=sheet_name)

@app.route('/get_cost/<product_name>')
def get_cost(product_name):
    df_inventory = pd.read_excel(EXCEL_FILE, sheet_name='inventory')
    row = df_inventory[df_inventory['Product_Name'] == product_name]

    if not row.empty:
        cost = float(row.iloc[0]['Selling_Price'])  # or 'Cost_per_unit' if purchase
        return jsonify({'cost': cost})
    return jsonify({'cost': 0})


@app.route('/sheets/<sheet_name>/delete/<int:row_index>', methods=['GET'])
def delete_row(sheet_name, row_index):
    try:
        # Read all sheets at once
        all_sheets = pd.read_excel(EXCEL_FILE, sheet_name=None)
        df = all_sheets.get(sheet_name)

        if df is None:
            flash('Sheet not found!', 'danger')
            return redirect(url_for('view_sheet', sheet_name=sheet_name))

        # Convert 1-based index to 0-based index
        actual_index = row_index - 1
        
        if actual_index < 0 or actual_index >= len(df):
            flash('Invalid row index!', 'danger')
            return redirect(url_for('view_sheet', sheet_name=sheet_name))

        # Get the row to be deleted before removing it
        row_to_delete = df.iloc[actual_index].copy()
        print(f"Deleting row: {row_to_delete.to_dict()}")  # Debug print

        # Update inventory if this is a purchase or sale
        if sheet_name.lower() in ['purchases', 'sales']:
            inventory_df = all_sheets.get('inventory')
            if inventory_df is not None:
                try:
                    print("Current inventory state:")
                    print(inventory_df[['HSN_Code', 'Product_Name', 'Units']].to_string())
                    
                    # Clean up HSN codes and product names for comparison
                    inventory_df['HSN_Code'] = inventory_df['HSN_Code'].astype(str).str.strip()
                    inventory_df['Product_Name'] = inventory_df['Product_Name'].astype(str).str.strip()
                    
                    # Get the values we're looking for
                    target_hsn = str(row_to_delete['HSN_Code']).strip()
                    target_product = str(row_to_delete['Product_Name']).strip()
                    print(f"Looking for HSN: '{target_hsn}', Product: '{target_product}'")
                    
                    # Find matching inventory row
                    inv_match = (
                        (inventory_df['HSN_Code'] == target_hsn) &
                        (inventory_df['Product_Name'] == target_product)
                    )
                    
                    print("Matching rows:")
                    print(inventory_df[inv_match][['HSN_Code', 'Product_Name', 'Units']].to_string())
                    
                    if inv_match.any():
                        inv_index = inv_match.idxmax()
                        current_stock = int(inventory_df.at[inv_index, 'Units'])
                        units = int(row_to_delete['Units'])
                        print(f"Current stock: {current_stock}, Units to adjust: {units}")

                        if sheet_name.lower() == 'purchases':
                            # For purchases: subtract units when deleting
                            new_stock = current_stock - units
                            inventory_df.at[inv_index, 'Units'] = new_stock
                            print(f"Purchase deletion: {current_stock} - {units} = {new_stock}")
                        elif sheet_name.lower() == 'sales':
                            # For sales: add units back when deleting
                            new_stock = current_stock + units
                            inventory_df.at[inv_index, 'Units'] = new_stock
                            print(f"Sale deletion: {current_stock} + {units} = {new_stock}")

                        print(f"Inventory updated: {units} units, new stock: {inventory_df.at[inv_index, 'Units']}")
                        all_sheets['inventory'] = inventory_df
                    else:
                        print("No matching inventory row found")
                        print("Available HSN codes:", inventory_df['HSN_Code'].unique())
                        print("Available Products:", inventory_df['Product_Name'].unique())
                except Exception as e:
                    print(f"Error updating inventory: {e}")
                    flash('Error updating inventory!', 'danger')
                    return redirect(url_for('view_sheet', sheet_name=sheet_name))

        # Remove the row from the sheet
        df = df.drop(df.index[actual_index]).reset_index(drop=True)
        all_sheets[sheet_name] = df

        # Save all sheets back
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='w') as writer:
            for name, data in all_sheets.items():
                data.to_excel(writer, sheet_name=name, index=False)

        flash('Row deleted successfully!', 'success')
        return redirect(url_for('view_sheet', sheet_name=sheet_name))

    except Exception as e:
        print(f"Error deleting row: {e}")
        flash('Error deleting row!', 'danger')
        return redirect(url_for('view_sheet', sheet_name=sheet_name))


if __name__ == '__main__':
    initialize_excel_file(EXCEL_FILE)
    app.run(debug=True)
